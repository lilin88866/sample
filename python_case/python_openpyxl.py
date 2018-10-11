from openpyxl import load_workbook
# wb = load_workbook('labdevice.xlsx')
# load_workbook('labdevice.xlsx')['TestENV'] 
# sheet_ranges = wb['TestENV'] 
sheet_dict={row[0].value:row[1].value for row in load_workbook('labdevice.xlsx')['TestENV'].rows if row[0] !=None}
print sheet_dict

print [(i,j) for i in [1,2,3] for j in [2,3,4]]