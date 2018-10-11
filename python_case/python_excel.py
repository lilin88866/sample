# use xlrd to read excel
import openpyxl
import xlrd
import xlwt

class OperExcel():
    #read Excel sheet
    def rExcel(self,inEfile,outfile):
        rfile = xlrd.open_workbook(inEfile)
        # create index and obtain an worksheet orderly
        table = rfile.sheet_by_index(0)
        #any way create
        # table = rfile.sheets()[0]
        # table = rfile.sheet_by_name(u'Sheet1')

        #obtain the hole row value and colum value
        table.row_values(0)
        table.col_values(0)

        #obtain the row num and colum num
        nrows = table.nrows - 1
        ncols = table.ncols
        print nrows #1
        print ncols #2
        
        # write value in cell
        # type 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
        table.put_cell(3, 1, 1, 'lilin', 0)
        print table.cell(3,1)
        print table.cell(3,1).value
        table.cell(0,0).value
        
        #obtain the list value looply
        for i in range(nrows):
            for j in range(ncols):
                # if i == j :
                    print table.cell(i,j).value
                    print table.row_values(i)
                    print table.col_values(j)
        # write the cell value into another file
        wfile = open(outfile,'w')
        #obtain all the first colum value
        for i in range(1,2):
            # print i
            for j in range(1,2):
                # print j
                table.cell(i,i).value="a"       #obtain the frist cell value 
                wfile.write(table.cell(i,j).value.encode('utf8') + '\n')
                wfile.write(table.cell(i,i).value)
        wfile.close()
  
# write value into Excel
    # def wExcel(self,infile,outEfile):
        # rfile = open(infile,'r')
        # buf = rfile.read().split('\n')
        # rfile.close()

        # w = xlwt.Workbook()
        # sheet = w.add_sheet('sheet1')
        # for i in range(len(buf)):
            # print buf[i]
            # sheet.write(i,0,buf[i].decode('utf8'))
            # w.save(outEfile)

if __name__ == '__main__':
    t = OperExcel()
    t.rExcel('test.xlsx','test1.xlsx')
    # t.wExcel('test','1.xls')


